# 2023-10-06 Initiated a py file
# We are including the Dividend in this calculation. Remember, this dividend is from Zerodha too. Because of that, this is slightly incorrect. But we cannoy do anything.

# import
import configparser
import os
from datetime import datetime
import openpyxl
import locale
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Set the locale to the Indian English format
locale.setlocale(locale.LC_ALL, 'en_IN')

# import - Own utils
from utility import SFUtil

# read configs
config = configparser.ConfigParser()        # instance
configFile = SFUtil.initiate_env_var('ConfigFile')
config.read(configFile)

def read_excel(file_path, pRow, pCol):
    data = sheet.cell(row=pRow, column=pCol).value

    return data

def sumUpExpenses(search_values):
    Total =0
    for row in sheet.iter_rows(values_only=True):
        cell_value_b = row[1]  # Value from column B (index 1)
        cell_value_e = row[4]  # Value from column E (index 4)
        if cell_value_b and any(search_value in str(cell_value_b) for search_value in search_values):
            # print(cell_value_b + str(cell_value_e))
            Total += cell_value_e or 0
    return Total

def sumUpCredits(search_values):
    Total =0
    for row in sheet.iter_rows(values_only=True):
        cell_value_b = row[1]  # Value from column B (index 1)
        cell_value_f = row[5]  # Value from column F (index 5)
        if cell_value_b and any(search_value in str(cell_value_b) for search_value in search_values):
            # print(cell_value_b + str(cell_value_e))
            Total += cell_value_f or 0
    return Total

intend = '      '
current_filename = os.path.basename(__file__)

print('['+datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'] [' + current_filename  + "] Started - Considering config: ["
      + configFile + "]")
print('Note: Exit Code 0 means Success')

# Variables assigning [common variables]
sourcePath = config['SelfFinance']['sourcepath']
filePrefix = config['SelfFinance']['statementprefix'] # sample file name: HDFC_BANK_STMT_FY2022 23.xlsx
grandFlow = 0
Years = 0

# Setting up the search variables
# eval is used to convert the string in to List
# C406Rent_search_values = eval(config['variable']['C406Rent_search_values'])
HDFCSEC_transfers = eval(config['variable']['HDFCSEC_transfers'])


# Get a list of files in the specified directory
files = os.listdir(sourcePath)

# Filter files that start with 'HDFC'
hdfcBank_files = [file for file in files if file.startswith(filePrefix)]

print(hdfcBank_files)

# Print the list of files
for file in hdfcBank_files:
    netflow = 0
    Years += 1
    file_path = os.path.join(sourcePath, file)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # Assuming you want to read from the active sheet

    # Get the total number of rows
    #total_rows = sheet.max_row

    result = read_excel(file_path, 6, 1)
    #print('Name: ' + result[:])

    result = read_excel(file_path, 16, 1)
    print(result[:])

    Total = sumUpExpenses(HDFCSEC_transfers)
    print('Debit  - HSEC Transfers: ' + locale.currency(Total, grouping=True))
    netflow = netflow + (-1 * Total)

    Total = sumUpCredits(HDFCSEC_transfers)
    print('Credit - HSEC Transfers: ' + locale.currency(Total, grouping=True))
    netflow = netflow + Total

    print(netflow)
    grandFlow = grandFlow + netflow
    print("")

print('Invested: ' + str(round(grandFlow, 2)) + ' INR in ' + str(Years) + ' years.')

# CurrentValue = input('CurrentValue: ')

